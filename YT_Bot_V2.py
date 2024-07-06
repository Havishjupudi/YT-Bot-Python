import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
import pyautogui
import time
import subprocess
import openpyxl

# Function to handle game selection
def on_game_selected():
    global selected_game
    selected_game = selected_var.get()
    update_radiobutton_styles()  # Update radiobutton styles when game is selected
    root.quit()

# Function to add game entries
def add_game():
    try:
        for i in range(6):
            row_frame = tk.Frame(radio_frame, bg="#212121")
            row_frame.pack(anchor="w", padx=10, pady=5)
            
            row_label = ttk.Label(row_frame, text=f"Game {i+1}: ", style="TLabel")
            row_label.pack(side="left")
            
            name_entry = tk.Entry(row_frame, font=("Helvetica", 12))
            name_entry.pack(side="left", padx=(0, 10))
            name_entries.append(name_entry)
            
            part_number_entry = tk.Entry(row_frame, font=("Helvetica", 12))
            part_number_entry.pack(side="left")
            part_number_entries.append(part_number_entry)
    except Exception as e:
        print("Error:", e)

# Function to save entries to Excel
def save_entries():
    try:
        wb = load_workbook("C:\\Users\\Havi\\OneDrive\\Desktop\\YT_Bot Project\\game_data.xlsx")
        sheet = wb.active

        for i in range(len(name_entries)):
            new_game_name = name_entries[i].get().strip()
            part_number = part_number_entries[i].get().strip()
            
            if new_game_name.lower() == "del" or part_number.lower() == "del":
                sheet.cell(row=i+2, column=1).value = None
                sheet.cell(row=i+2, column=2).value = None
            else:
                if new_game_name:
                    sheet.cell(row=i+2, column=1).value = new_game_name
                if part_number:
                    try:
                        part_number_int = int(part_number)
                        sheet.cell(row=i+2, column=2).value = part_number_int
                    except ValueError:
                        print(f"Invalid part number: {part_number} at row {i+2}")

         # Select your Excel path Correctly
        wb.save("C:\\Users\\Havi\\OneDrive\\Desktop\\YT_Bot Project\\game_data.xlsx") 
        print("Changes saved successfully.")
    except Exception as e:
        print("Error:", e)

# Load game data from Excel
def load_game_data():
    # Select your Excel path Correctly
    wb = load_workbook("C:\\Users\\Havi\\OneDrive\\Desktop\\YT_Bot Project\\game_data.xlsx")
    sheet = wb.active
    game_data = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)]
    return game_data

# Function to update radiobutton styles
def update_radiobutton_styles():
    selected_value = selected_var.get()
    for rb in radio_buttons:
        if rb['text'] == selected_value:
            rb.configure(style="Selected.TRadiobutton")
        else:
            rb.configure(style="TRadiobutton")

# Initialize selected game and entry lists
selected_game = None
name_entries = []
part_number_entries = []

# Create main window
root = tk.Tk()
root.title("Select Game")
root.configure(background="black")  # Set background color of the root window to black

# Configure themed style
style = ttk.Style(root)
style.configure("TFrame", background="black")  # Set background color for frames
style.configure("TRadiobutton", background="black", foreground="#ffffff", font=("Helvetica", 12))  # Set font, background, and foreground color for radiobuttons
style.configure("TLabel", background="black", foreground="#007bff", font=("Helvetica", 14, "bold"))  # Set font, background, and foreground color for labels
style.configure("TButton", foreground="#ffffff", font=("Helvetica", 12, "bold"), background="#28a745")  # Set font, background, and foreground color for buttons
style.map("TRadiobutton", background=[("selected", "#007bff")])  # Set background color when radiobutton is selected

# Load game data from Excel
game_names = load_game_data()

# Create frame for radiobuttons
radio_frame = ttk.Frame(root)
radio_frame.pack(pady=10)

# Initialize selected variable
selected_var = tk.StringVar()

# Create radiobuttons for each game
radio_buttons = []
for game in game_names:
    rb = ttk.Radiobutton(radio_frame, text=game, variable=selected_var, value=game)
    rb.pack(anchor="w", padx=20, pady=5)
    radio_buttons.append(rb)

# Create button to add game names
add_button = tk.Button(root, text="Add Game", command=add_game, bg='orange', fg='white', font=('Arial', 12, 'bold'))
add_button.pack(side=tk.LEFT, anchor=tk.NW, padx=(10, 5), pady=(10, 0))

# Create button to save entries
save_button = tk.Button(root, text="Save Entries", command=save_entries, bg='red', fg='white', font=('Arial', 12, 'bold'))
save_button.pack(side=tk.RIGHT, anchor=tk.NW, padx=(5, 10), pady=(10, 0))

# Create button to confirm selection
confirm_button = tk.Button(root, text="Confirm", command=on_game_selected, bg='green', fg='white', font=('Arial', 12, 'bold'))
confirm_button.pack(pady=10)

# Start GUI
root.mainloop()

# After GUI is closed, selected_game will contain the selected game name
if selected_game:
    print("Selected Game:", selected_game)
    game_name = selected_game
    print("Now using selected game:", game_name)
else:
    print("No game selected.")

# Additional functions for handling image recognition and subprocess commands
def locate_image(image_path, confidence=0.8):
    image_paths = f"C:\\Users\\Havi\\OneDrive\\Desktop\\YT_Bot Project\\{image_path}"
    while True:
        try:
            img = pyautogui.locateOnScreen(image_paths, confidence=confidence)
            if img is not None:
                return img
            else:
                print(f"{image_path} not found, retrying...")
                time.sleep(1)
        except pyautogui.ImageNotFoundException:
            print(f"Exception: {image_path} not found, retrying...")
            time.sleep(1)

def read_game_data(game_data_file):
    wb = openpyxl.load_workbook(game_data_file)
    sheet = wb.active
    game_data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        game_name = row[0].strip().lower() if row[0] else ""
        part_number = row[1]
        game_data[game_name] = part_number
    return wb, sheet, game_data

def get_part_number(game_data, game_name):
    game_name = game_name.strip().lower()
    part_number = game_data.get(game_name)
    return part_number

def update_part_number(wb, sheet, game_name, new_part_number):
    game_name = game_name.strip().lower()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        if row[0].value and row[0].value.strip().lower() == game_name:
            row[1].value = new_part_number
            break
    wb.save(game_data_file)

game_data_file = "C:\\Users\\Havi\\OneDrive\\Desktop\\YT_Bot Project\\game_data.xlsx"
wb, sheet, game_data = read_game_data(game_data_file)

part_no = get_part_number(game_data, game_name)

if part_no is None:
    print(f"Part number not found for {game_name} in the Excel file.")
    exit()

time.sleep(1)

command = r'"C:\Program Files\Google\Chrome\Application\chrome_proxy.exe" --profile-directory="Profile 1" --app-id=cahggfghendlbihgniaflhickgjcohcb --start-fullscreen'
subprocess.Popen(command, shell=True)
time.sleep(5)

dismiss_img = locate_image('dismiss.png', 0.7) # Select your IMG path Correctly
pyautogui.click(dismiss_img)
time.sleep(10)

edit_img = locate_image('edit.png', 0.9) # Select your IMG path Correctly
pyautogui.click(edit_img)
time.sleep(1)

title_img = locate_image('title_req.png', 0.9) # Select your IMG path Correctly
pyautogui.click(title_img)
time.sleep(0.5)

pyautogui.hotkey('ctrl', 'a')
pyautogui.press('delete')

pyautogui.write(f'{game_name} | PC Walkthrough | Part-{str(part_no)}')
time.sleep(0.5)

pyautogui.scroll(-1000)
time.sleep(1)
pyautogui.click(x=918, y=450)

change_img = locate_image('change.png', 0.9) # Select your IMG path Correctly
pyautogui.click(change_img)
time.sleep(0.5)

path_img = locate_image('IMG path.png', 0.9) # Select your IMG path Correctly
pyautogui.click(path_img)
time.sleep(0.5)

pyautogui.write(f'C:\\Users\\Havi\\OneDrive\\Pictures\\Thumbnails\\{game_name}')
pyautogui.press('enter')
time.sleep(0.5)

search_box_img = locate_image('search_box.png', 0.9) # Select your IMG path Correctly
pyautogui.click(search_box_img)
time.sleep(0.5)

pyautogui.write(f'{part_no}.png')
pyautogui.press('enter')

select_img = locate_image('select.png', 0.9) # Select your IMG path Correctly
pyautogui.click(select_img)
time.sleep(0.5)

if part_no == 1:
    new_pl_img = locate_image('new_pl.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(new_pl_img)
    time.sleep(1)

    new_img = locate_image('new.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(new_img)
    time.sleep(1)

    title_img = locate_image('title_req.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(title_img)
    time.sleep(1)

    pyautogui.write(game_name)

    create_img = locate_image('create.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(create_img)
    time.sleep(1)
else:
    search_pl_img = locate_image('search_pl.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(search_pl_img)
    time.sleep(0.5)

    pyautogui.write(game_name)
    time.sleep(0.5)

    box_img = locate_image('box.png', 0.9) # Select your IMG path Correctly
    pyautogui.click(box_img)
    time.sleep(0.5)

done_img = locate_image('done.png', 0.9) # Select your IMG path Correctly
pyautogui.click(done_img)
time.sleep(0.5)

save_img = locate_image('save.png', 0.9) # Select your IMG path Correctly
pyautogui.click(save_img)
time.sleep(0.5)

pyautogui.hotkey('f11')

# OBS STUDIO

obs_img = locate_image('obs.png', 0.9) # Select your IMG path Correctly
pyautogui.click(obs_img)
time.sleep(0.5)

scene_img = locate_image('scene.png', 0.9) # Select your IMG path Correctly
pyautogui.click(scene_img)
time.sleep(0.5)

mgbc_img = locate_image('mgbc.png', 0.9) # Select your IMG path Correctly
pyautogui.click(mgbc_img)
time.sleep(0.5)

select_eb_img = locate_image('select_eb.png', 0.9) # Select your IMG path Correctly
pyautogui.click(select_eb_img)
time.sleep(0.5)

auto_img = locate_image('auto.png', 0.9) # Select your IMG path Correctly
pyautogui.click(auto_img)
time.sleep(0.5)

sbc_img = locate_image('sbc.png', 0.9) # Select your IMG path Correctly
pyautogui.click(sbc_img)
time.sleep(2)

pyautogui.click(x=1803, y=13)
time.sleep(0.3)

pyautogui.click(x=1803, y=13)
time.sleep(0.3)

# Increment part number by 1
new_part_no = part_no + 1
update_part_number(wb, sheet, game_name, new_part_no)

print(f"Part Number incremented to: {new_part_no}")
